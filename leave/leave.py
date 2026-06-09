#!/usr/bin/env python3
"""
Duha International School — Staff Leave Ledger  v2 (Self-Configuring)
  Config sheet     → add/remove staff, leave types, set year here
  Leave Records    → add new leave entries here (dropdowns auto-fill details)
  Leave Summary    → auto-calculated per-staff totals & balances  (READ ONLY)
  Monthly Breakdown→ auto-calculated month-wise breakdown         (READ ONLY)
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date

# ═══════════════════════════════════════════════════════════
#  PALETTE
# ═══════════════════════════════════════════════════════════
NAVY  ="1F3864"; BLUE  ="2E75B6"; DBLUE ="17375E"
GOLD  ="BF8F00"; LGOLD ="FFF2CC"; LGREEN="E2EFDA"
ALT   ="EAF2FF"; WHITE ="FFFFFF"; LGREY ="F2F2F2"
AUTO  ="DDEEFF"                        # auto-calculated cell bg
INPUT ="FFFAEE"                        # user-input cell bg
GF="C6EFCE"; GT="276221"
RF="FFC7CE"; RT="9C0006"
OF="FFEB9C"; OT="7F3F00"
TOTBG ="17375E"; WARN="FFFDE7"

# ═══════════════════════════════════════════════════════════
#  STYLE HELPERS
# ═══════════════════════════════════════════════════════════
def Fl(h): return PatternFill("solid",start_color=h,end_color=h)
def Bd(c="BBBBBB"):
    s=Side(border_style="thin",color=c)
    return Border(left=s,right=s,top=s,bottom=s)
def MBd(c="555555"):
    s=Side(border_style="medium",color=c)
    return Border(left=s,right=s,top=s,bottom=s)
def Ft(bold=False,sz=9,col="000000",it=False):
    return Font(name="Arial",bold=bold,size=sz,color=col,italic=it)
def Al(h="center",v="center",wrap=True):
    return Alignment(horizontal=h,vertical=v,wrap_text=wrap)

def title_row(ws,row,text,nc,sz=13,bg=NAVY):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=nc)
    c=ws.cell(row,1,text)
    c.fill=Fl(bg); c.font=Ft(True,sz,WHITE); c.alignment=Al(); c.border=MBd()
    ws.row_dimensions[row].height=28

def note_row(ws,row,text,nc,bg=WARN,fg="7F5200",italic=True):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=nc)
    c=ws.cell(row,1,text)
    c.fill=Fl(bg); c.font=Ft(False,8,fg,it=italic); c.alignment=Al("left"); c.border=Bd(GOLD)
    ws.row_dimensions[row].height=17

def readonly_row(ws,row,text,nc):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=nc)
    c=ws.cell(row,1,text)
    c.fill=Fl("FBE4D5"); c.font=Ft(True,8,"9C0006"); c.alignment=Al("center"); c.border=MBd("9C0006")
    ws.row_dimensions[row].height=17

def sec_hdr(ws,row,sc,ec,text,bg=GOLD):
    if sc!=ec: ws.merge_cells(start_row=row,start_column=sc,end_row=row,end_column=ec)
    c=ws.cell(row,sc,text)
    c.fill=Fl(bg); c.font=Ft(True,10,WHITE); c.alignment=Al(); c.border=MBd(GOLD)
    ws.row_dimensions[row].height=22

def col_hdr(ws,row,col,text,bg=DBLUE):
    c=ws.cell(row,col,text)
    c.fill=Fl(bg); c.font=Ft(True,9,WHITE); c.alignment=Al(); c.border=MBd()
    return c

def mhdr(ws,row,sc,ec,text,bg=DBLUE):
    if sc!=ec: ws.merge_cells(start_row=row,start_column=sc,end_row=row,end_column=ec)
    col_hdr(ws,row,sc,text,bg)

def stat_cell(cell,status):
    fills={"Approved":(GF,GT),"Pending":(OF,OT),"Cancelled":(RF,RT)}
    bg,fg=fills.get(status,(WHITE,"000000"))
    cell.value=status; cell.fill=Fl(bg); cell.font=Ft(True,9,fg)
    cell.alignment=Al(); cell.border=Bd()

def set_cols(ws,widths):
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width=w

# ═══════════════════════════════════════════════════════════
#  DATA
# ═══════════════════════════════════════════════════════════
REPORT_YEAR=2026

STAFF=[
    ("Hasan Mahmud","Director"),
    ("Hasina Mohammed","Principal"),
    ("Nusrat Jahan Ira",""),
    ("Sadia Sultana Rima","Preschool Co-ordinator"),
    ("Fatema Akter Mili",""),
    ("Farida Yesmin Tamanna",""),
    ("Shalina Akter Eity",""),
    ("Rabia Sultana Prity","Class Teacher"),
    ("Afrin Jahan Uzma","Lead Teacher"),
    ("Ariful Islam Maruf",""),
    ("Jinat Farhana Chy","Asst. Co-ord. (Pre-School)"),
    ("Tabassuma Barat Chowdhury","Primary Co-ordinator"),
    ("Sumaya Rahman","Asst. Co-ord. (Secondary)"),
    ("Warda Siraj","Head Of BC"),
    ("Irfan Hossain",""),
    ("Mohoshina Shifa Buble","Teacher's Assistant"),
    ("Sadia Akter Lubna","Teacher's Assistant"),
    ("Taslima Akter","Subject Teacher (IC)"),
    ("Samiya Hashem","Teacher's Assistant"),
    ("Humaira Tasnim",""),
    ("Aziza Sultana","Teacher's Assistant"),
    ("Farhana Yesmin",""),
    ("Kamrun Nahar","Class Teacher"),
    ("Tashmem Al Faed","Subject Teacher"),
    ("Md Murad Hoshen","IC Co-ordinator"),
    ("Najmul Haque Masum","Subject Teacher (IC)"),
    ("Rahnuman Nesa Chowdhury","Class Teacher"),
    ("Othora Nausin Jafrin","Assistant Teacher"),
    ("Md Rahat Chy",""),
    ("Sania Rahman","Class Teacher"),
    ("Jaharin Subah","Class Teacher"),
    ("Afroza Akter",""),
    ("Ummay Hani Juthi","Lead Teacher"),
    ("Tanzina Haque","Subject Teacher"),
    ("Tahsina Tarannum",""),
    ("Mohammad Arafat Ul Alam",""),
    ("Md Kawser Hussain Rasel",""),
    ("Md. Mainuddin",""),
    ("Sumaiya Nusrat Pushpo","Assistant Teacher"),
    ("Ismat Tasnia","Assistant Teacher"),
    ("Ayesha Siddika Ruba","Subject Teacher"),
    ("Jannatur Rahman",""),
    ("Sultana Yasmin Geeti",""),
    ("Shaila Parvin",""),
    ("Ashik Bhuiyan",""),
    ("Md Helal Uddin",""),
    ("Shahida",""),
    ("Rabeya Rima",""),
    ("Moyna",""),
    ("Rohima",""),
    ("Sharmin",""),
    ("Nargis Akter",""),
    ("Sanjida Islam",""),
]

LEAVE_TYPES=[("Casual Leave",10),("Special Leave",5)]

# (name, type, reason, from, to, days, req_date, status)
REC=[
    # March 2026 CL entries
    ("Hasina Mohammed","Casual Leave","March CL",date(2026,3,1),date(2026,3,2),2,date(2026,3,1),"Approved"),
    ("Farida Yesmin Tamanna","Casual Leave","March CL",date(2026,3,1),date(2026,3,2),2,date(2026,3,1),"Approved"),
    ("Rabia Sultana Prity","Casual Leave","March CL",date(2026,3,1),date(2026,3,5),5,date(2026,3,1),"Approved"),
    ("Afrin Jahan Uzma","Casual Leave","March CL",date(2026,3,1),date(2026,3,1),1,date(2026,3,1),"Approved"),
    ("Ariful Islam Maruf","Casual Leave","March CL",date(2026,3,1),date(2026,3,2),2,date(2026,3,1),"Approved"),
    ("Jinat Farhana Chy","Casual Leave","March CL",date(2026,3,1),date(2026,3,1),1,date(2026,3,1),"Approved"),
    ("Irfan Hossain","Casual Leave","March CL",date(2026,3,1),date(2026,3,4),4,date(2026,3,1),"Approved"),
    ("Mohoshina Shifa Buble","Casual Leave","March CL",date(2026,3,1),date(2026,3,3),3,date(2026,3,1),"Approved"),
    ("Taslima Akter","Casual Leave","March CL",date(2026,3,1),date(2026,3,2),2,date(2026,3,1),"Approved"),
    ("Humaira Tasnim","Casual Leave","March CL",date(2026,3,1),date(2026,3,7),7,date(2026,3,1),"Approved"),
    ("Aziza Sultana","Casual Leave","March CL",date(2026,3,1),date(2026,3,2),2,date(2026,3,1),"Approved"),
    ("Farhana Yesmin","Casual Leave","March CL",date(2026,3,1),date(2026,3,1),1,date(2026,3,1),"Approved"),
    ("Kamrun Nahar","Casual Leave","March CL",date(2026,3,1),date(2026,3,5),5,date(2026,3,1),"Approved"),
    ("Najmul Haque Masum","Casual Leave","March CL",date(2026,3,1),date(2026,3,2),2,date(2026,3,1),"Approved"),
    ("Othora Nausin Jafrin","Casual Leave","March CL",date(2026,3,1),date(2026,3,1),1,date(2026,3,1),"Approved"),
    ("Ummay Hani Juthi","Casual Leave","March CL",date(2026,3,1),date(2026,3,3),3,date(2026,3,1),"Approved"),
    ("Tanzina Haque","Casual Leave","March CL",date(2026,3,1),date(2026,3,1),1,date(2026,3,1),"Approved"),
    ("Tahsina Tarannum","Casual Leave","March CL",date(2026,3,1),date(2026,3,1),1,date(2026,3,1),"Approved"),
    ("Mohammad Arafat Ul Alam","Casual Leave","March CL",date(2026,3,1),date(2026,3,3),3,date(2026,3,1),"Approved"),
    ("Md. Mainuddin","Casual Leave","March CL",date(2026,3,1),date(2026,3,2),2,date(2026,3,1),"Approved"),
    ("Jannatur Rahman","Casual Leave","March CL",date(2026,3,1),date(2026,3,1),1,date(2026,3,1),"Approved"),
    ("Shaila Parvin","Casual Leave","March CL",date(2026,3,1),date(2026,3,1),1,date(2026,3,1),"Approved"),
    ("Sanjida Islam","Casual Leave","March CL",date(2026,3,1),date(2026,3,14),14,date(2026,3,1),"Approved"),

    # Existing April / May / June entries
    ("Hasina Mohammed","Casual Leave","Appointment",date(2026,5,5),date(2026,5,5),1,date(2026,4,28),"Cancelled"),
    ("Sadia Sultana Rima","Special Leave","Family emergency",date(2026,5,18),date(2026,5,18),1,date(2026,5,19),"Approved"),
    ("Sadia Sultana Rima","Casual Leave","-",date(2026,4,19),date(2026,4,19),1,date(2026,4,26),"Approved"),
    ("Sadia Sultana Rima","Casual Leave","-",date(2026,4,23),date(2026,4,23),1,date(2026,4,26),"Approved"),
    ("Sadia Sultana Rima","Casual Leave","-",date(2026,4,19),date(2026,4,19),1,date(2026,4,26),"Cancelled"),
    ("Rabia Sultana Prity","Casual Leave","MA final examination",date(2026,5,11),date(2026,5,11),1,date(2026,5,7),"Cancelled"),
    ("Afrin Jahan Uzma","Casual Leave","Health condition / fever",date(2026,5,6),date(2026,5,7),2,date(2026,5,6),"Approved"),
    ("Afrin Jahan Uzma","Casual Leave","Endoscopy & medical tests",date(2026,5,4),date(2026,5,4),1,date(2026,5,5),"Approved"),
    ("Afrin Jahan Uzma","Casual Leave","-",date(2026,4,16),date(2026,4,16),1,date(2026,4,26),"Approved"),
    ("Jinat Farhana Chy","Casual Leave","-",date(2026,4,23),date(2026,4,26),4,date(2026,4,26),"Approved"),
    ("Tabassuma Barat Chowdhury","Special Leave","Family program",date(2026,5,7),date(2026,5,7),1,date(2026,5,7),"Pending"),
    ("Sumaya Rahman","Casual Leave","Family need",date(2026,5,13),date(2026,5,13),1,date(2026,5,13),"Approved"),
    ("Warda Siraj","Special Leave","Early leave",date(2026,4,30),date(2026,4,30),1,date(2026,4,30),"Approved"),
    ("Warda Siraj","Casual Leave","Test",date(2026,4,28),date(2026,4,28),1,date(2026,4,28),"Cancelled"),
    ("Warda Siraj","Casual Leave","-",date(2026,4,25),date(2026,4,25),1,date(2026,4,26),"Approved"),
    ("Mohoshina Shifa Buble","Casual Leave","Fever and Low BP",date(2026,5,11),date(2026,5,11),1,date(2026,5,11),"Approved"),
    ("Mohoshina Shifa Buble","Casual Leave","Absent on 17th counted as CL",date(2026,5,17),date(2026,5,17),1,date(2026,5,17),"Approved"),
    ("Mohoshina Shifa Buble","Casual Leave","Absent on 21st counted as CL",date(2026,5,21),date(2026,5,21),1,date(2026,5,21),"Approved"),
    ("Sadia Akter Lubna","Casual Leave","Illness",date(2026,5,6),date(2026,5,6),1,date(2026,5,6),"Approved"),
    ("Sadia Akter Lubna","Casual Leave","Absent on 10th counted as CL",date(2026,5,10),date(2026,5,10),1,date(2026,5,10),"Approved"),
    ("Sadia Akter Lubna","Casual Leave","Absent on 11th counted as CL",date(2026,5,11),date(2026,5,11),1,date(2026,5,11),"Approved"),
    ("Sadia Akter Lubna","Casual Leave","High fever / tonsillitis",date(2026,4,28),date(2026,4,28),1,date(2026,4,28),"Approved"),
    ("Taslima Akter","Casual Leave","Severe sickness",date(2026,4,18),date(2026,4,23),5,date(2026,4,26),"Approved"),
    ("Samiya Hashem","Casual Leave","Outside Chittagong (urgent)",date(2026,5,6),date(2026,5,6),1,date(2026,5,6),"Approved"),
    ("Aziza Sultana","Special Leave","Sick (early leave 11:10 AM)",date(2026,5,20),date(2026,5,20),1,date(2026,5,20),"Pending"),
    ("Farhana Yesmin","Casual Leave","Absent on 3rd counted as CL",date(2026,5,3),date(2026,5,3),1,date(2026,5,3),"Approved"),
    ("Kamrun Nahar","Casual Leave","Absent on 14th counted as CL",date(2026,5,14),date(2026,5,14),1,date(2026,5,14),"Approved"),
    ("Kamrun Nahar","Casual Leave","Absent on 19th counted as CL",date(2026,5,19),date(2026,5,19),1,date(2026,5,19),"Approved"),
    ("Kamrun Nahar","Casual Leave","Exam",date(2026,4,27),date(2026,4,27),1,date(2026,4,26),"Approved"),
    ("Kamrun Nahar","Casual Leave","Master's Exam (early leave)",date(2026,4,26),date(2026,4,26),1,date(2026,4,29),"Pending"),
    ("Kamrun Nahar","Casual Leave","Half day (both days)",date(2026,4,19),date(2026,4,20),2,date(2026,4,26),"Approved"),
    ("Tashmem Al Faed","Casual Leave","Severe headache / cold / blurry vision",date(2026,5,17),date(2026,5,17),1,date(2026,5,17),"Approved"),
    ("Md Murad Hoshen","Casual Leave","-",date(2026,4,9),date(2026,4,10),2,date(2026,4,26),"Approved"),
    ("Najmul Haque Masum","Special Leave","Migraine (early leave)",date(2026,5,13),date(2026,5,13),1,date(2026,5,13),"Pending"),
    ("Rahnuman Nesa Chowdhury","Casual Leave","Flight delayed (overseas)",date(2026,6,6),date(2026,6,6),1,date(2026,6,5),"Pending"),
    ("Rahnuman Nesa Chowdhury","Casual Leave","Eye infection",date(2026,5,17),date(2026,5,17),1,date(2026,5,17),"Approved"),
    ("Othora Nausin Jafrin","Casual Leave","-",date(2026,4,25),date(2026,4,26),2,date(2026,4,26),"Approved"),
    ("Sania Rahman","Casual Leave","Mother sick",date(2026,5,12),date(2026,5,12),1,date(2026,5,12),"Approved"),
    ("Jaharin Subah","Casual Leave","-",date(2026,4,25),date(2026,4,25),1,date(2026,4,26),"Approved"),
    ("Ummay Hani Juthi","Casual Leave","Injured hip (accidental fall)",date(2026,5,14),date(2026,5,14),1,date(2026,5,14),"Approved"),
    ("Tanzina Haque","Casual Leave","Food poisoning",date(2026,5,9),date(2026,5,9),1,date(2026,5,21),"Cancelled"),
    ("Tanzina Haque","Special Leave","Sick (after taking all classes)",date(2026,5,17),date(2026,5,17),1,date(2026,5,17),"Approved"),
    ("Sumaiya Nusrat Pushpo","Casual Leave","Sickness",date(2026,5,10),date(2026,5,10),1,date(2026,5,10),"Approved"),
    ("Sumaiya Nusrat Pushpo","Casual Leave","Sickness",date(2026,5,6),date(2026,5,6),1,date(2026,5,5),"Approved"),
    ("Sumaiya Nusrat Pushpo","Casual Leave","Sickness",date(2026,5,5),date(2026,5,5),1,date(2026,5,5),"Approved"),
    ("Sumaiya Nusrat Pushpo","Casual Leave","Absent on 12th counted as CL",date(2026,5,12),date(2026,5,12),1,date(2026,5,12),"Approved"),
    ("Ismat Tasnia","Special Leave","Personal emergency",date(2026,5,13),date(2026,5,13),1,date(2026,5,13),"Approved"),
    ("Ayesha Siddika Ruba","Special Leave","Family issue",date(2026,4,13),date(2026,4,13),1,date(2026,5,13),"Approved"),
    ("Ayesha Siddika Ruba","Casual Leave","Sickness",date(2026,4,26),date(2026,4,27),2,date(2026,4,28),"Approved"),
    ("Ayesha Siddika Ruba","Casual Leave","-",date(2026,4,15),date(2026,4,15),1,date(2026,4,26),"Approved"),
    ("Sultana Yasmin Geeti","Casual Leave","Absent on 12th counted as CL",date(2026,5,12),date(2026,5,12),1,date(2026,5,12),"Approved"),
]

# ═══════════════════════════════════════════════════════════
#  CELL ADDRESS CONSTANTS
#  Config layout (9 cols A-I):
#    Staff   → B7:B106 (name), C7:C106 (designation)  — 100 slots
#    LvTypes → E7:E11  (type),  F7:F11  (annual days) — 5  slots
#    Settings→ H7:I9   (setting name, value)
#       I7 = Report Year
# ═══════════════════════════════════════════════════════════
CFG_STAFF_START = 7
CFG_STAFF_END   = 106          # 100 slots
CFG_LT_START    = 7
CFG_LT_END      = 11           # 5 slots
CFG_YEAR_CELL   = "Config!$I$7"
CFG_VLOOKUP     = "Config!$B$7:$C$106"
CFG_STAFF_DROP  = "Config!$B$7:$B$106"
CFG_LT_DROP     = "Config!$E$7:$E$11"
LR              = "'Leave Records'"   # sheet reference in formulas
NC_CFG=9; NC_REC=11; NC_SUM=24; NC_MON=16

MONTH_NAMES=["Jan","Feb","Mar","Apr","May","Jun",
             "Jul","Aug","Sep","Oct","Nov","Dec"]

# ═══════════════════════════════════════════════════════════
# FORMULA BUILDERS
# ═══════════════════════════════════════════════════════════
def f_desig(r):
    return f'=IFERROR(VLOOKUP(B{r},{CFG_VLOOKUP},2,0),"")'

def f_month(r):
    return f'=IF(F{r}="","",TEXT(F{r},"MMMM YYYY"))'

def f_sl_rec(r):
    return f'=IF(B{r}="","",ROW()-3)'

def f_sl_sum(r):
    return f'=IF(B{r}="","",ROW()-4)'

def f_staff_name(r):
    cfg_r = r + 2          # Summary row 5 → Config row 7
    return f'=IF(Config!$B${cfg_r}="","",Config!$B${cfg_r})'

def f_staff_desig(r):
    cfg_r = r + 2
    return f'=IF(Config!$B${cfg_r}="","",Config!$C${cfg_r})'

def f_lt_allot(r, lt_cfg_row):
    return f'=IF($B{r}="","",IFERROR(Config!$F${lt_cfg_row},0))'

def f_lt_stat(r, lt_cfg_row, status):
    return (f'=IF($B{r}="","",IF(Config!$E${lt_cfg_row}="",0,'
            f'SUMIFS({LR}!$I:$I,{LR}!$B:$B,$B{r},'
            f'{LR}!$D:$D,Config!$E${lt_cfg_row},'
            f'{LR}!$K:$K,"{status}")))')

def f_lt_balance(r, allot_col, apprv_col):
    ac=get_column_letter(allot_col); pc=get_column_letter(apprv_col)
    return f'=IF($B{r}="","",{ac}{r}-{pc}{r})'

def f_total_approved(r):
    # sum all 5 "Approved" cols: E,I,M,Q,U  (cols 5,9,13,17,21)
    approved_cols="E,I,M,Q,U"
    parts="+".join(f"{c}{r}" for c in approved_cols.split(","))
    return f'=IF($B{r},"",{parts})' if False else f'=IF($B{r}="","",{parts})'

def f_monthly(r, month_num):
    return (f'=IF($B{r}="","",SUMIFS({LR}!$I:$I,{LR}!$B:$B,$B{r},'
            f'{LR}!$K:$K,"Approved",'
            f'{LR}!$F:$F,">="&DATE({CFG_YEAR_CELL},{month_num},1),'
            f'{LR}!$F:$F,"<="&EOMONTH(DATE({CFG_YEAR_CELL},{month_num},1),0)))')

def f_year_total(r):
    return f'=IF($B{r}="","",SUM(D{r}:O{r}))'

# ═══════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ═══════════════════════════════════════════════════════════
wb = Workbook()

# ─────────────────────────────────────────────────────────
#  SHEET 1 — Config
# ─────────────────────────────────────────────────────────
ws_c = wb.active
ws_c.title = "Config"

title_row(ws_c,1,"DUHA INTERNATIONAL SCHOOL — CONFIGURATION PANEL",NC_CFG,bg=GOLD)
note_row(ws_c,2,
    "INSTRUCTIONS: (1) Staff Roster — add/remove staff rows freely. "
    "(2) Leave Types — add/remove leave types and set annual days. Max 5 types. "
    "(3) Settings — change Report Year only. "
    "Blue cells are auto-filled. Yellow cells are editable.",
    NC_CFG, bg=LGOLD, fg="5C3317")

# Blank spacer row
ws_c.row_dimensions[3].height = 6

# Row 4 — Section headers
sec_hdr(ws_c,4,1,3,"STAFF ROSTER")
ws_c.cell(4,4).fill = Fl(WHITE)
sec_hdr(ws_c,4,5,6,"LEAVE TYPES")
ws_c.cell(4,7).fill = Fl(WHITE)
sec_hdr(ws_c,4,8,9,"SETTINGS",bg=DBLUE)

# Row 5 — Section instructions
for sc,ec,txt,bg in [
    (1,3,"Add or remove staff rows. Delete row content to remove a staff member.",LGOLD),
    (5,6,"Add/remove leave types. Annual Days = days allowed per year.",LGOLD),
    (8,9,"Change Report Year to switch annual view in Monthly Breakdown.",WARN)
]:
    ws_c.merge_cells(start_row=5,start_column=sc,end_row=5,end_column=ec)
    c=ws_c.cell(5,sc,txt)
    c.fill=Fl(bg); c.font=Ft(False,8,"5C3317",it=True)
    c.alignment=Al("left",wrap=True); c.border=Bd(GOLD)
ws_c.row_dimensions[5].height=20

# Row 6 — Column headers
for col,txt,bg in [
    (1,"#",DBLUE),(2,"Staff Name",BLUE),(3,"Designation",BLUE),
    (5,"Leave Type Name",BLUE),(6,"Annual Days",BLUE),
    (8,"Setting",DBLUE),(9,"Value",DBLUE)
]:
    col_hdr(ws_c,6,col,txt,bg)
ws_c.cell(6,4).border=Bd(); ws_c.cell(6,7).border=Bd()
ws_c.row_dimensions[6].height=22

# Staff data rows 7-106
for i,(nm,desig) in enumerate(STAFF):
    r=i+7; alt=(i%2==0); bg=ALT if alt else WHITE
    # Auto #
    c=ws_c.cell(r,1,f"=IF(B{r}=\"\",\"\",ROW()-6)")
    c.fill=Fl(AUTO); c.font=Ft(sz=9); c.alignment=Al(); c.border=Bd()
    # Name (yellow = user editable)
    c=ws_c.cell(r,2,nm)
    c.fill=Fl(INPUT); c.font=Ft(bold=True,sz=9); c.alignment=Al("left"); c.border=Bd()
    # Designation
    c=ws_c.cell(r,3,desig)
    c.fill=Fl(INPUT); c.font=Ft(sz=9); c.alignment=Al("left"); c.border=Bd()
    ws_c.row_dimensions[r].height=17

# Blank staff slots (rows after current staff up to 106)
for r in range(len(STAFF)+7, CFG_STAFF_END+1):
    c=ws_c.cell(r,1,f"=IF(B{r}=\"\",\"\",ROW()-6)")
    c.fill=Fl(AUTO); c.font=Ft(sz=9); c.alignment=Al(); c.border=Bd()
    for col in (2,3):
        c=ws_c.cell(r,col)
        c.fill=Fl(LGREY); c.border=Bd()
    ws_c.row_dimensions[r].height=15

# "Add row" hint below staff
ws_c.merge_cells(start_row=CFG_STAFF_END+1,start_column=1,end_row=CFG_STAFF_END+1,end_column=3)
c=ws_c.cell(CFG_STAFF_END+1,1,"↑ Add new staff in rows above. Maximum 100.")
c.fill=Fl(LGOLD); c.font=Ft(False,8,"5C3317",it=True)
c.alignment=Al("center"); c.border=Bd(GOLD)

# Leave types (rows 7-11, cols E-F)
for i,(lt,days) in enumerate(LEAVE_TYPES):
    r=CFG_LT_START+i
    c=ws_c.cell(r,5,lt)
    c.fill=Fl(INPUT); c.font=Ft(bold=True,sz=9); c.alignment=Al("left"); c.border=Bd()
    c=ws_c.cell(r,6,days)
    c.fill=Fl(INPUT); c.font=Ft(sz=9); c.alignment=Al("center"); c.border=Bd()

for r in range(CFG_LT_START+len(LEAVE_TYPES), CFG_LT_END+1):
    for col in (5,6):
        c=ws_c.cell(r,col); c.fill=Fl(LGREY); c.border=Bd()
    ws_c.row_dimensions[r].height=15

ws_c.merge_cells(start_row=CFG_LT_END+1,start_column=5,end_row=CFG_LT_END+1,end_column=6)
c=ws_c.cell(CFG_LT_END+1,5,"↑ Max 5 leave types.")
c.fill=Fl(LGOLD); c.font=Ft(False,8,"5C3317",it=True)
c.alignment=Al("center"); c.border=Bd(GOLD)

# Settings (cols H-I, rows 7+)
settings=[("Report Year",REPORT_YEAR),("School Name","Duha International School"),
          ("Prepared By","Principal / Admin")]
for i,(sname,sval) in enumerate(settings):
    r=7+i
    c=ws_c.cell(r,8,sname)
    c.fill=Fl(LGREY); c.font=Ft(bold=True,sz=9); c.alignment=Al("left"); c.border=Bd()
    c=ws_c.cell(r,9,sval)
    # Year is editable (yellow), others are info
    c.fill=Fl(LGOLD if i==0 else LGREY)
    c.font=Ft(bold=(i==0),sz=9); c.alignment=Al("center"); c.border=Bd()

# Data validation on Config: Annual Days must be 1-365
days_dv=DataValidation(type="whole",operator="between",formula1="1",formula2="365",
    showErrorMessage=True,error="Enter a number between 1 and 365.",errorTitle="Invalid")
ws_c.add_data_validation(days_dv)
days_dv.sqref=f"F{CFG_LT_START}:F{CFG_LT_END}"

# Data validation on Config: Year must be reasonable
year_dv=DataValidation(type="whole",operator="between",formula1="2020",formula2="2035",
    showErrorMessage=True,error="Enter a year between 2020 and 2035.",errorTitle="Invalid Year")
ws_c.add_data_validation(year_dv)
year_dv.sqref="I7"

set_cols(ws_c,[5,28,26,2,22,10,2,18,12])
ws_c.sheet_properties.tabColor="BF8F00"
ws_c.freeze_panes="A7"

# ─────────────────────────────────────────────────────────
#  SHEET 2 — Leave Records
#  Cols: A=SL# B=Name C=Desig D=Type E=Reason F=From G=To H=Month I=Days J=ReqDate K=Status
#  Data rows 4-303  (300 rows pre-formatted)
# ─────────────────────────────────────────────────────────
ws_r = wb.create_sheet("Leave Records")

title_row(ws_r,1,"DUHA INTERNATIONAL SCHOOL — LEAVE RECORDS",NC_REC,bg=BLUE)
note_row(ws_r,2,
    "FILL IN (white cells): Staff Name ▼  Leave Type ▼  Reason  From Date  "
    "To Date  Days  Request Date  Status ▼     |     "
    "AUTO-FILLED (blue cells): SL#  Designation  Month",
    NC_REC, bg="E9F5FF", fg="17375E")

# Column headers row 3
hdr_data=["SL#","Staff Name","Designation","Leave Type","Reason",
          "From Date","To Date","Month","Days","Request Date","Status"]
hdr_bg=[AUTO,BLUE,AUTO,BLUE,BLUE,BLUE,BLUE,AUTO,BLUE,BLUE,BLUE]
for ci,(h,bg) in enumerate(zip(hdr_data,hdr_bg),1):
    col_hdr(ws_r,3,ci,h,bg)
ws_r.row_dimensions[3].height=35
ws_r.freeze_panes="A4"
ws_r.auto_filter.ref=f"A3:{get_column_letter(NC_REC)}3"

# Data validation
name_dv=DataValidation(type="list",formula1=CFG_STAFF_DROP,allow_blank=True,
    showErrorMessage=True,error="Select from the dropdown.",errorTitle="Invalid Name")
type_dv=DataValidation(type="list",formula1=CFG_LT_DROP,allow_blank=True,
    showErrorMessage=True,error="Select from the dropdown.",errorTitle="Invalid Type")
stat_dv=DataValidation(type="list",formula1='"Approved,Pending,Cancelled"',allow_blank=True,
    showErrorMessage=True,error="Choose: Approved, Pending, or Cancelled.",errorTitle="Invalid Status")
for dv in (name_dv,type_dv,stat_dv): ws_r.add_data_validation(dv)
name_dv.sqref="B4:B303"
type_dv.sqref="D4:D303"
stat_dv.sqref="K4:K303"

# Write existing records (rows 4-45)
DESIG={nm:d for nm,d in STAFF}
for i,(nm,lt,rsn,fd,td,dy,rd,st) in enumerate(REC):
    r=i+4; alt=(i%2==0); bg=ALT if alt else WHITE
    # A: SL# (formula)
    c=ws_r.cell(r,1,f_sl_rec(r))
    c.fill=Fl(AUTO); c.font=Ft(); c.alignment=Al(); c.border=Bd()
    # B: Name (user input)
    c=ws_r.cell(r,2,nm)
    c.fill=Fl(INPUT); c.font=Ft(bold=True); c.alignment=Al("left"); c.border=Bd()
    # C: Designation (formula)
    c=ws_r.cell(r,3,f_desig(r))
    c.fill=Fl(AUTO); c.font=Ft(); c.alignment=Al("left"); c.border=Bd()
    # D: Leave Type
    c=ws_r.cell(r,4,lt)
    c.fill=Fl(INPUT); c.font=Ft(); c.alignment=Al(); c.border=Bd()
    # E: Reason
    c=ws_r.cell(r,5,rsn)
    c.fill=Fl(INPUT); c.font=Ft(); c.alignment=Al("left"); c.border=Bd()
    # F: From Date
    c=ws_r.cell(r,6,fd)
    c.fill=Fl(INPUT); c.font=Ft(); c.alignment=Al(); c.border=Bd(); c.number_format="DD-MMM-YYYY"
    # G: To Date
    c=ws_r.cell(r,7,td)
    c.fill=Fl(INPUT); c.font=Ft(); c.alignment=Al(); c.border=Bd(); c.number_format="DD-MMM-YYYY"
    # H: Month (formula)
    c=ws_r.cell(r,8,f_month(r))
    c.fill=Fl(AUTO); c.font=Ft(); c.alignment=Al(); c.border=Bd()
    # I: Days
    c=ws_r.cell(r,9,dy)
    c.fill=Fl(INPUT); c.font=Ft(bold=True); c.alignment=Al(); c.border=Bd()
    # J: Request Date
    c=ws_r.cell(r,10,rd)
    c.fill=Fl(INPUT); c.font=Ft(); c.alignment=Al(); c.border=Bd(); c.number_format="DD-MMM-YYYY"
    # K: Status (coloured)
    stat_cell(ws_r.cell(r,11),st)
    ws_r.row_dimensions[r].height=17

# Pre-format blank rows 46-303 with formulas
for r in range(len(REC)+4, 304):
    alt=((r-4)%2==0); bg=ALT if alt else WHITE
    # auto cols: A,C,H
    for col,formula in [(1,f_sl_rec(r)),(3,f_desig(r)),(8,f_month(r))]:
        c=ws_r.cell(r,col,formula)
        c.fill=Fl(AUTO); c.font=Ft(); c.alignment=Al("left" if col==3 else "center"); c.border=Bd()
    # user input cols: B,D,E,F,G,I,J,K
    for col in (2,4,5,6,7,9,10,11):
        c=ws_r.cell(r,col)
        c.fill=Fl(bg); c.font=Ft(); c.border=Bd()
        if col in (6,7,10): c.number_format="DD-MMM-YYYY"
    ws_r.row_dimensions[r].height=15

# "Add rows here" note
ws_r.merge_cells(start_row=304,start_column=1,end_row=304,end_column=NC_REC)
c=ws_r.cell(304,1,"← 300 rows pre-formatted above. Insert rows before this line to add more.")
c.fill=Fl(LGOLD); c.font=Ft(False,8,"5C3317",it=True)
c.alignment=Al("center"); c.border=Bd(GOLD)

set_cols(ws_r,[5,26,24,14,36,13,13,14,7,13,12])
ws_r.sheet_properties.tabColor="2E75B6"

# ─────────────────────────────────────────────────────────
#  SHEET 3 — Leave Summary
#  Cols: A=SL B=Name C=Desig | D-G=LT1 | H-K=LT2 | L-O=LT3 | P-S=LT4 | T-W=LT5 | X=Total
#  LT group cols (per type): Allot | Approved | Pending | Balance
#  Staff rows: 5-104 (100 rows), mapping: summary_row+2 = config_row
# ─────────────────────────────────────────────────────────
ws_s = wb.create_sheet("Leave Summary")

title_row(ws_s,1,"DUHA INTERNATIONAL SCHOOL — LEAVE SUMMARY",NC_SUM)
readonly_row(ws_s,2,
    "READ ONLY — Auto-calculated from 'Leave Records'. "
    "To add staff: edit Config sheet.  To add leave entries: edit Leave Records sheet.",NC_SUM)

# Row 3 — Group headers
mhdr(ws_s,3,1,1,"SL")
mhdr(ws_s,3,2,2,"Staff Name")
mhdr(ws_s,3,3,3,"Designation")
# LT group headers (dynamic: pulls name from Config)
for i in range(5):
    sc_=4+i*4; ec_=7+i*4; cfg_r=CFG_LT_START+i
    ws_s.merge_cells(start_row=3,start_column=sc_,end_row=3,end_column=ec_)
    c=ws_s.cell(3,sc_,
        f'=IFERROR(IF(Config!$E${cfg_r}="","— Not Set —",Config!$E${cfg_r}),"— Not Set —")')
    c.fill=Fl(DBLUE); c.font=Ft(True,9,WHITE); c.alignment=Al(); c.border=MBd()
mhdr(ws_s,3,24,24,"Grand Total\nApproved")
ws_s.row_dimensions[3].height=28

# Row 4 — Column headers
col_hdr(ws_s,4,1,"SL")
col_hdr(ws_s,4,2,"Staff Name")
col_hdr(ws_s,4,3,"Designation")
for i in range(5):
    base=4+i*4
    for off,txt in [(0,"Annual\nAllot."),(1,"Days\nApproved"),(2,"Days\nPending"),(3,"Balance\n(Allot-Apprv)")]:
        bg=BLUE if off!=3 else "17375E"
        col_hdr(ws_s,4,base+off,txt,bg)
col_hdr(ws_s,4,24,"Total\nApproved",DBLUE)
ws_s.row_dimensions[4].height=40
ws_s.freeze_panes="D5"
ws_s.auto_filter.ref=f"A4:{get_column_letter(NC_SUM)}4"

# Data rows 5-104 (100 staff rows)
for i in range(100):
    r=i+5; alt=(i%2==0); bg=ALT if alt else WHITE
    # SL
    c=ws_s.cell(r,1,f_sl_sum(r)); c.fill=Fl(AUTO); c.font=Ft(); c.alignment=Al(); c.border=Bd()
    # Name + Desig from Config
    for col,formula in [(2,f_staff_name(r)),(3,f_staff_desig(r))]:
        c=ws_s.cell(r,col,formula)
        c.fill=Fl(bg); c.font=Ft(bold=(col==2)); c.alignment=Al("left"); c.border=Bd()
    # 5 LT groups
    for lt_i in range(5):
        lt_cfg_row=CFG_LT_START+lt_i
        base_col=4+lt_i*4
        allot_col=base_col; apprv_col=base_col+1; pend_col=base_col+2; bal_col=base_col+3
        cells_and_formulas=[
            (allot_col, f_lt_allot(r,lt_cfg_row), LGREY),
            (apprv_col, f_lt_stat(r,lt_cfg_row,"Approved"), GF if i%2==0 else "E8F5E9"),
            (pend_col,  f_lt_stat(r,lt_cfg_row,"Pending"),  OF if i%2==0 else "FFFDE7"),
            (bal_col,   f_lt_balance(r,allot_col,apprv_col), bg),
        ]
        for col,formula,cell_bg in cells_and_formulas:
            c=ws_s.cell(r,col,formula)
            c.fill=Fl(cell_bg); c.font=Ft(); c.alignment=Al(); c.border=Bd()
    # Grand Total Approved = sum of all 5 "Approved" cols
    apprv_cols=[get_column_letter(5+lt_i*4) for lt_i in range(5)]
    total_f=f'=IF($B{r}="","",{"+".join(f"{c}{r}" for c in apprv_cols)})'
    c=ws_s.cell(r,24,total_f)
    c.fill=Fl(LGOLD if alt else "FFF8E1"); c.font=Ft(bold=True); c.alignment=Al(); c.border=MBd(GOLD)
    ws_s.row_dimensions[r].height=17

# Totals row
tr=105
ws_s.merge_cells(start_row=tr,start_column=1,end_row=tr,end_column=3)
c=ws_s.cell(tr,1,"SCHOOL TOTALS")
c.fill=Fl(TOTBG); c.font=Ft(True,9,WHITE); c.alignment=Al(); c.border=MBd()
for col in range(4,NC_SUM+1):
    cl=get_column_letter(col)
    c=ws_s.cell(tr,col,f"=SUM({cl}5:{cl}104)")
    c.fill=Fl(TOTBG); c.font=Ft(True,9,WHITE); c.alignment=Al(); c.border=MBd()
ws_s.row_dimensions[tr].height=20

set_cols(ws_s,[5,26,24,9,11,10,10, 9,11,10,10, 9,11,10,10, 9,11,10,10, 9,11,10,10, 13])
ws_s.sheet_properties.tabColor="1F3864"

# ─────────────────────────────────────────────────────────
#  SHEET 4 — Monthly Breakdown
#  Cols: A=SL B=Name C=Desig D=Jan E=Feb ... O=Dec P=YearTotal
#  Year is dynamic: reads from Config!I7
# ─────────────────────────────────────────────────────────
ws_m = wb.create_sheet("Monthly Breakdown")

title_row(ws_m,1,"DUHA INTERNATIONAL SCHOOL — MONTHLY LEAVE BREAKDOWN",NC_MON)
readonly_row(ws_m,2,
    "READ ONLY — Approved leave days per month, all types combined. "
    f"Report Year is set in Config sheet (cell I7). Currently: {REPORT_YEAR}  |  "
    "Change Config!I7 to update all month formulas automatically.",NC_MON)

# Row 3 — Group headers
mhdr(ws_m,3,1,3,"Staff")
mhdr(ws_m,3,4,15,f"Monthly Approved Leave Days  (Year auto-reads from Config!I7)")
mhdr(ws_m,3,16,16,"Year\nTotal")
ws_m.row_dimensions[3].height=25

# Row 4 — Month column headers
col_hdr(ws_m,4,1,"SL"); col_hdr(ws_m,4,2,"Staff Name"); col_hdr(ws_m,4,3,"Designation")
for mi,mn in enumerate(MONTH_NAMES):
    col_hdr(ws_m,4,4+mi,mn,bg=BLUE if mi%2==0 else DBLUE)
col_hdr(ws_m,4,16,"Year\nTotal",DBLUE)
ws_m.row_dimensions[4].height=30
ws_m.freeze_panes="D5"

# Data rows 5-104
for i in range(100):
    r=i+5; alt=(i%2==0); bg=ALT if alt else WHITE
    # SL
    c=ws_m.cell(r,1,f_sl_sum(r)); c.fill=Fl(AUTO); c.font=Ft(); c.alignment=Al(); c.border=Bd()
    # Name + Desig
    for col,formula in [(2,f_staff_name(r)),(3,f_staff_desig(r))]:
        c=ws_m.cell(r,col,formula)
        c.fill=Fl(bg); c.font=Ft(bold=(col==2)); c.alignment=Al("left"); c.border=Bd()
    # 12 monthly formulas (cols D-O)
    month_fills=["EAF2FF","FFFFFF","EAF2FF","C6EFCE","C6EFCE","EAF2FF",
                 "EAF2FF","FFFFFF","EAF2FF","C6EFCE","C6EFCE","EAF2FF"]
    for mi in range(12):
        c=ws_m.cell(r,4+mi,f_monthly(r,mi+1))
        c.fill=Fl(month_fills[mi] if alt else WHITE)
        c.font=Ft(); c.alignment=Al(); c.border=Bd()
    # Year Total
    c=ws_m.cell(r,16,f_year_total(r))
    c.fill=Fl(LGOLD if alt else "FFF8E1"); c.font=Ft(bold=True); c.alignment=Al(); c.border=MBd(GOLD)
    ws_m.row_dimensions[r].height=17

# Totals row
tr=105
ws_m.merge_cells(start_row=tr,start_column=1,end_row=tr,end_column=3)
c=ws_m.cell(tr,1,"SCHOOL TOTALS")
c.fill=Fl(TOTBG); c.font=Ft(True,9,WHITE); c.alignment=Al(); c.border=MBd()
for col in range(4,NC_MON+1):
    cl=get_column_letter(col)
    c=ws_m.cell(tr,col,f"=SUM({cl}5:{cl}104)")
    c.fill=Fl(TOTBG); c.font=Ft(True,9,WHITE); c.alignment=Al(); c.border=MBd()
ws_m.row_dimensions[tr].height=20

set_cols(ws_m,[5,26,22,8,8,8,8,8,8,8,8,8,8,8,8,11])
ws_m.sheet_properties.tabColor="375623"

# ─────────────────────────────────────────────────────────
#  FINALIZE
# ─────────────────────────────────────────────────────────
wb._sheets = [ws_c, ws_r, ws_s, ws_m]

# Save to the current directory where the script resides
script_dir = os.path.dirname(os.path.abspath(__file__))
out = os.path.join(script_dir, "Duha_Leave_Ledger_v2_Configurable.xlsx")
wb.save(out)
print(f"Saved: {out}")
print(f"Sheets: {wb.sheetnames}")
print(f"Staff: {len(STAFF)} | LT: {len(LEAVE_TYPES)} | Records: {len(REC)}")